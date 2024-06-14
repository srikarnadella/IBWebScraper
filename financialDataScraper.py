import yfinance as yf  # Library for fetching financial data from Yahoo Finance
import pandas as pd    # Library for data manipulation and analysis
import numpy as np     # Library for numerical operations
import argparse       # Library for parsing command-line arguments
from openpyxl import Workbook     # Library for creating Excel files
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill  # Styles and formatting in Excel
from openpyxl.utils.dataframe import dataframe_to_rows    # Convert Pandas DataFrame to Excel
from openpyxl.chart import LineChart, Reference           # Creating charts in Excel
from openpyxl.formatting.rule import ColorScaleRule       # Conditional formatting in Excel
from datetime import datetime   # Date and time manipulation in Python


def get_financial_data(ticker):
    """
    Fetches financial data for a given stock ticker using Yahoo Finance.

    Parameters:
    ticker (str): Ticker symbol of the company.

    Returns:
    financials (DataFrame): DataFrame containing financial statements.
    balance_sheet (DataFrame): DataFrame containing balance sheet data.
    cash_flow (DataFrame): DataFrame containing cash flow statement data.
    """
    try:
        stock = yf.Ticker(ticker)
        financials = stock.financials       # Fetching financial statements
        balance_sheet = stock.balance_sheet  # Fetching balance sheet
        cash_flow = stock.cashflow           # Fetching cash flow statements
        return financials, balance_sheet, cash_flow
    except Exception as e:
        print(f"Failed to retrieve data for ticker: {ticker}. Error: {e}")
        return None, None, None


def parse_financial_data(financials, balance_sheet, cash_flow):
    """
    Parses and transposes financial data if available.

    Parameters:
    financials (DataFrame): DataFrame containing financial statements.
    balance_sheet (DataFrame): DataFrame containing balance sheet data.
    cash_flow (DataFrame): DataFrame containing cash flow statement data.

    Returns:
    financials (DataFrame): Transposed financial statements DataFrame.
    balance_sheet (DataFrame): Transposed balance sheet DataFrame.
    cash_flow (DataFrame): Transposed cash flow statement DataFrame.
    """
    if financials is not None and not financials.empty:
        financials = financials.T   # Transpose financials for easier analysis
    else:
        print("No financial data to parse.")
        return None, None, None

    if balance_sheet is not None and not balance_sheet.empty:
        balance_sheet = balance_sheet.T   # Transpose balance sheet
    else:
        print("No balance sheet data to parse.")
        return None, None, None

    if cash_flow is not None and not cash_flow.empty:
        cash_flow = cash_flow.T   # Transpose cash flow statement
    else:
        print("No cash flow data to parse.")
        return None, None, None

    return financials, balance_sheet, cash_flow


def estimate_growth_rate(cash_flow):
    """
    Estimates the average growth rate of free cash flows.

    Parameters:
    cash_flow (DataFrame): DataFrame containing cash flow statement data.

    Returns:
    growth_rate (float): Estimated average growth rate of free cash flows.
    """
    if cash_flow is None:
        return 0.02  # Default growth rate if cash flow data is missing

    free_cash_flows = cash_flow.loc[:, 'Free Cash Flow'].dropna().astype(float)  # Extracting free cash flows
    if len(free_cash_flows) < 2:
        return 0.02  # Default growth rate if insufficient data

    growth_rates = free_cash_flows.pct_change().dropna()  # Calculate percentage change
    return np.mean(growth_rates)  # Average growth rate


def calculate_dcf(free_cash_flows, discount_rate, growth_rate, years):
    """
    Calculates the Discounted Cash Flow (DCF) value of a company.

    Parameters:
    free_cash_flows (list): List of free cash flows over multiple periods.
    discount_rate (float): Discount rate (weighted average cost of capital).
    growth_rate (float): Estimated growth rate of free cash flows.
    years (int): Number of years for DCF calculation.

    Returns:
    dcf (float): Calculated DCF value of the company.
    """
    dcf = 0
    for i in range(len(free_cash_flows)):
        # Discounting each year's cash flow
        dcf += free_cash_flows[i] / ((1 + discount_rate) ** (i + 1))

    # Calculating terminal value using perpetual growth model
    terminal_value = free_cash_flows[-1] * (1 + growth_rate) / (discount_rate - growth_rate)
    dcf += terminal_value / ((1 + discount_rate) ** years)  # Discounting terminal value

    return dcf


def create_styles(wb):
    """
    Creates and adds custom styles for Excel cells.

    Parameters:
    wb (Workbook): Excel workbook object.
    """
    # Create named styles if they don't already exist
    if "currency" not in wb.named_styles:
        currency_format = NamedStyle(name="currency")
        currency_format.number_format = "$#,##0.00"
        wb.add_named_style(currency_format)

    if "percentage" not in wb.named_styles:
        percentage_format = NamedStyle(name="percentage")
        percentage_format.number_format = "0.00%"
        wb.add_named_style(percentage_format)

    if "general" not in wb.named_styles:
        general_format = NamedStyle(name="general")
        general_format.number_format = "#,##0"
        wb.add_named_style(general_format)


def style_worksheet(ws, wb):
    """
    Applies styles and formatting to Excel worksheet.

    Parameters:
    ws (Worksheet): Excel worksheet object.
    wb (Workbook): Excel workbook object.
    """
    header_font = Font(bold=True)

    # Apply header font and alignment to the first row and first column
    for cell in ws['A'] + ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Determine column widths based on cell content length
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if isinstance(cell.value, (int, float)):
                # Apply specific styles for different types of financial metrics
                if 'Revenue' in ws.cell(row=1, column=cell.column).value or 'Income' in ws.cell(row=1, column=cell.column).value:
                    cell.style = "currency"
                elif 'Margin' in ws.cell(row=1, column=cell.column).value or 'Rate' in ws.cell(row=1, column=cell.column).value:
                    cell.style = "percentage"
                else:
                    cell.style = "general"

            try:
                # Determine the maximum content length in each column for setting column width
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width  # Set column width


def add_chart(ws, title, data_range, cats_range):
    """
    Adds a line chart to the Excel worksheet.

    Parameters:
    ws (Worksheet): Excel worksheet object.
    title (str): Title of the chart.
    data_range (list): Range of data to plot [min_col, min_row, max_col, max_row].
    cats_range (list): Range of categories [min_col, min_row, max_col, max_row].
    """
    chart = LineChart()  # Create a line chart object
    chart.title = title  # Set chart title
    chart.style = 13     # Set chart style
    chart.y_axis.title = 'Value'  # Set y-axis title
    chart.x_axis.title = 'Date'   # Set x-axis title

    # Define references for data and categories
    data = Reference(ws, min_col=data_range[0], min_row=data_range[1], max_col=data_range[2], max_row=data_range[3])
    cats = Reference(ws, min_col=cats_range[0], min_row=cats_range[1], max_col=cats_range[2], max_row=cats_range[3])

    chart.add_data(data, titles_from_data=True)  # Add data to the chart
    chart.set_categories(cats)  # Set chart categories

    ws.add_chart(chart, "E5")   # Add chart to the worksheet at specific coordinates


def extract_year_from_date(date):
    """
    Extracts the year from a datetime object.

    Parameters:
    date (Timestamp): Pandas Timestamp object representing a date.

    Returns:
    int: Year extracted from the date.
    """
    if isinstance(date, pd.Timestamp):
        return date.year
    return None


def save_to_excel(financial_data, balance_sheet, cash_flow, dcf_value, ticker):
    """
    Saves financial data and DCF calculation results to an Excel file.

    Parameters:
    financial_data (DataFrame): DataFrame containing financial data.
    balance_sheet (DataFrame): DataFrame containing balance sheet data.
    cash_flow (DataFrame): DataFrame containing cash flow statement data.
    dcf_value (float): Calculated Discounted Cash Flow (DCF) value.
    ticker (str): Ticker symbol of the company.
    """
    filename = f"{ticker}_financial_data.xlsx"
    wb = Workbook()
    
    # Create styles
    create_styles(wb)
    
    # Financial Data Sheet
    ws1 = wb.active
    ws1.title = "Financial Data"
    
    # Preprocess financial data to extract year from the date index
    financial_data['Year'] = financial_data.index.map(extract_year_from_date)
    financial_data = financial_data.reset_index(drop=True)
    
    # Convert the first column to just the years
    ws1.append(['Year'] + list(financial_data.columns)[1:])
    for row in financial_data.itertuples(index=False):
        ws1.append([row.Year] + list(row)[1:])

    style_worksheet(ws1, wb)
    add_chart(ws1, "Revenue and Net Income", [2, 1, len(financial_data) + 1, 3], [2, 2, len(financial_data) + 1, 2])
    
    # Balance Sheet Data Sheet
    ws2 = wb.create_sheet(title="Balance Sheet")
    for row in dataframe_to_rows(balance_sheet, index=True, header=True):
        ws2.append(row)
    style_worksheet(ws2, wb)
    
    # Cash Flow Data Sheet
    ws3 = wb.create_sheet(title="Cash Flow")
    for row in dataframe_to_rows(cash_flow, index=True, header=True):
        ws3.append(row)
    style_worksheet(ws3, wb)
    
    # DCF Calculation Sheet
    ws4 = wb.create_sheet(title="DCF Calculation")
    ws4.append(["Metric", "Value"])
    ws4.append(["DCF Value", dcf_value])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws4["B2"].number_format = "$#,##0.00"
    style_worksheet(ws4, wb)
    
    # Summary Sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Ticker", ticker])
    ws_summary.append(["DCF Value", dcf_value])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_summary["B2"].number_format = "$#,##0.00"
    style_worksheet(ws_summary, wb)

    # Adding conditional formatting to Summary sheet
    color_scale_rule = ColorScaleRule(start_type='min', start_color='FF0000', end_type='max', end_color='00FF00')
    ws_summary.conditional_formatting.add('B2:B2', color_scale_rule)
    
    wb.save(filename)


def main(ticker):
    """
    Fetches financial data, performs DCF valuation, and saves results to Excel.

    Parameters:
    ticker (str): Ticker symbol of the company.
    """
    financials, balance_sheet, cash_flow = get_financial_data(ticker)

    if financials is not None and balance_sheet is not None and cash_flow is not None:
        parsed_financials, parsed_balance_sheet, parsed_cash_flow = parse_financial_data(financials, balance_sheet, cash_flow)
        
        if parsed_financials is not None and parsed_balance_sheet is not None and parsed_cash_flow is not None:
            growth_rate = estimate_growth_rate(parsed_cash_flow)
            discount_rate = 0.10  # Example discount rate
            years = 5  # Example number of years

            free_cash_flows = parsed_cash_flow['Free Cash Flow'].dropna().astype(float).tolist()
            dcf_value = calculate_dcf(free_cash_flows, discount_rate, growth_rate, years)
            
            save_to_excel(parsed_financials, parsed_balance_sheet, parsed_cash_flow, dcf_value, ticker)
            print(f"DCF Value for {ticker}: ${dcf_value:.2f}")
        else:
            print("Failed to parse financial data.")
    else:
        print("Failed to retrieve financial data.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fetch financial data and calculate DCF for a given ticker.")
    parser.add_argument("ticker", type=str, nargs='?', help="Ticker symbol of the company.")
    args = parser.parse_args()

    if args.ticker:
        ticker = args.ticker
    else:
        ticker = input("Please enter the ticker symbol: ")

    main(ticker)
